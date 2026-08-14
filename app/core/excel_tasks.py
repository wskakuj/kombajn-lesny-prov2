"""
Kombajn Leśny PRO — Zadania Excel
==================================
Zależności: config.py (EXCEL_SHEET_DEFAULTS)
Odpowiada za: funkcje pomocnicze dla Excela — parsowanie właścicieli,
              łączenie plików XLS/VAL, wykonywanie makr VBA, formatowanie arkuszy.

Te funkcje są standalone (nie wymagają instancji ModernApp).
Można je testować w izolacji.
"""

import re
import warnings
import json
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter

from app.config import EXCEL_SHEET_DEFAULTS

# Wyciszanie ostrzeżeń z xlrd przy czytaniu starych .xls
warnings.filterwarnings("ignore", message=".*OLE2 inconsistency.*")
warnings.filterwarnings("ignore", message=".*file size.*not.*sector size.*")
warnings.filterwarnings("ignore", message=".*SSCS size.*")

def bezpieczna_liczba(val):
    if pd.isna(val) or val == '' or val == 'nan':
        return 0.0
    s = str(val).replace(',', '.')
    s = re.sub(r'\s+', '', s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def wczytaj_i_przetworz_wlascicieli(sciezka_do_pliku):
    # DYNAMICZNE SZUKANIE NAGŁÓWKA
    df_raw = pd.read_excel(sciezka_do_pliku, header=None, nrows=20)
    header_row = 1
    for i, row in df_raw.iterrows():
        row_str = " ".join([str(val).lower() for val in row.values])
        if 'numer działki' in row_str or 'numer dzialki' in row_str:
            header_row = i
            break

    df = pd.read_excel(sciezka_do_pliku, header=header_row)
    df = df.rename(columns={'Numer działki': 'nr_dz', 'Pow.\nklasouż.': 'Pow. klasouż.'})

    kolumny_do_wypelnienia = ['nr_dz', 'J. rej.', 'Pow. działki', 'Właściciel']
    istniejace_kolumny = [col for col in kolumny_do_wypelnienia if col in df.columns]

    for col in istniejace_kolumny:
        df[col] = df[col].astype(str).replace(r'^\s*$', np.nan, regex=True)
        df[col] = df[col].replace('nan', np.nan)
        df[col] = df[col].replace('-', np.nan)

    if istniejace_kolumny:
        df[istniejace_kolumny] = df[istniejace_kolumny].ffill()

    if 'nr_dz' in df.columns:
        df['nr_dz'] = df['nr_dz'].astype(str).str.strip()
        df['nr_dz'] = df['nr_dz'].str.replace(r'\.0$', '', regex=True)

    if 'J. rej.' in df.columns:
        def extract_after_g(val):
            v_str = str(val)
            if 'G' in v_str:
                return v_str.split('G')[-1]
            return v_str

        df['J. rej.'] = df['J. rej.'].apply(extract_after_g)

    if 'Właściciel' in df.columns:
        df['Właściciel'] = df['Właściciel'].astype(str).replace(['nan', 'NaN'], 'Brak danych')

    # --- DODANA LOGIKA POBIERANIA WŁAŚCICIELA DO GŁÓWNEJ PAMIĘCI ---
    cols_full = ['nr_dz', 'J. rej.', 'Pow. działki']
    if 'Właściciel' in df.columns:
        cols_full.append('Właściciel')

    df_full = df[cols_full].drop_duplicates(
        'nr_dz').copy() if 'nr_dz' in df.columns and 'Pow. działki' in df.columns else pd.DataFrame()
    # ---------------------------------------------------------------

    if not df_full.empty:
        df_full['pow dz'] = df_full['Pow. działki'].apply(bezpieczna_liczba)

    wiersze_po_rozbiciu = []
    for _, row in df.iterrows():
        klasy = str(row.get('Klasoużytek', '')).split('\n')
        powierzchnie = str(row.get('Pow. klasouż.', '')).split('\n')
        max_len = max(len(klasy), len(powierzchnie))
        klasy += [''] * (max_len - len(klasy))
        powierzchnie += [''] * (max_len - len(powierzchnie))
        for k, p in zip(klasy, powierzchnie):
            nowy_wiersz = row.copy()
            nowy_wiersz['Klasoużytek'] = k
            nowy_wiersz['Pow. klasouż.'] = p
            wiersze_po_rozbiciu.append(nowy_wiersz)

    df_exploded = pd.DataFrame(wiersze_po_rozbiciu)

    for col in ['Pow. działki', 'Pow. klasouż.']:
        if col in df_exploded.columns:
            df_exploded[col] = df_exploded[col].apply(bezpieczna_liczba)

    if 'Klasoużytek' in df_exploded.columns:
        df_ls = df_exploded[df_exploded['Klasoużytek'].astype(str).str.contains('Ls', case=False, na=False)].copy()
    else:
        df_ls = pd.DataFrame(columns=df_exploded.columns)

    if df_ls.empty:
        wynik = pd.DataFrame(columns=['nr_dz', 'J. rej.', 'pow dz', 'pow ls', 'Właściciel'])
        return wynik, df_full

    wynik = df_ls.groupby(['nr_dz', 'J. rej.', 'Pow. działki', 'Właściciel'], as_index=False)['Pow. klasouż.'].sum()
    wynik = wynik.rename(columns={'Pow. działki': 'pow dz', 'Pow. klasouż.': 'pow ls'})
    wynik['pow dz'] = wynik['pow dz'].round(4)
    wynik['pow ls'] = wynik['pow ls'].round(4)
    wynik = wynik[['nr_dz', 'J. rej.', 'pow dz', 'pow ls', 'Właściciel']]
    return wynik, df_full


def wczytaj_i_przetworz_val(sciezka_do_pliku_val):
    try:
        with open(sciezka_do_pliku_val, 'r', encoding='cp1250') as file:
            linie = file.readlines()
    except Exception as e:
        print(f"Błąd wczytywania VAL: {e}")
        return None

    dane_wyjsciowe = []
    aktualny_nr_dz = None

    for line in reversed(linie):
        line = line.strip()
        if not line or line.startswith(';'):
            continue
        elementy = line.split()
        if not elementy:
            continue
        if elementy[0] == '*':
            if len(elementy) >= 2:
                aktualny_nr_dz = elementy[1]
        elif elementy[0] == '^':
            if len(elementy) >= 3:
                oznaczenie = elementy[1]
                if 'X' not in oznaczenie and re.search(r'[A-Za-z]', oznaczenie):
                    litera = oznaczenie
                    if aktualny_nr_dz:
                        pow_sqm_str = elementy[2]
                        try:
                            pow_geo = float(pow_sqm_str.replace(',', '.')) / 10000.0
                        except ValueError:
                            continue
                        if pow_geo >= 0.001:
                            dane_wyjsciowe.append({
                                'nr_dz': aktualny_nr_dz,
                                'litera': litera,
                                'pow geo': round(pow_geo, 4)
                            })

    dane_wyjsciowe.reverse()
    df = pd.DataFrame(dane_wyjsciowe)
    if df.empty:
        return pd.DataFrame(columns=['nr_dz', 'litera', 'pow geo'])
    df = df[['nr_dz', 'litera', 'pow geo']]
    return df


def polacz_xls_i_val(df_xls, df_full, df_val):
    xls = df_xls.copy()
    val = df_val.copy()

    df_merged = pd.merge(val, xls, on='nr_dz', how='left')

    mapping_j_rej = df_full.set_index('nr_dz')['J. rej.']
    mapping_pow_dz = df_full.set_index('nr_dz')['pow dz']

    df_merged['J. rej.'] = df_merged['J. rej.'].fillna(df_merged['nr_dz'].map(mapping_j_rej))
    df_merged['pow dz'] = df_merged['pow dz'].fillna(df_merged['nr_dz'].map(mapping_pow_dz))

    # --- NOWA LOGIKA RATOWANIA NAZWISK DLA DZIAŁEK "PRZYBYŁO" ---
    if 'Właściciel' in df_full.columns:
        mapping_wlasciciel = df_full.set_index('nr_dz')['Właściciel']
        df_merged['Właściciel'] = df_merged['Właściciel'].fillna(df_merged['nr_dz'].map(mapping_wlasciciel))
    # ------------------------------------------------------------

    df_out = pd.DataFrame()
    df_out['Kolumna_A'] = ""
    df_out['J. rej.'] = df_merged['J. rej.']
    df_out['nr_dz'] = df_merged['nr_dz']
    df_out['litery'] = df_merged['litera']
    df_out['pow geo'] = df_merged['pow geo']
    df_out['TU POWSTANĄ DANE'] = np.nan
    df_out['Kolumna_G'] = ""
    df_out['nr_dz_ewid'] = df_merged['nr_dz']
    df_out['pow ls'] = df_merged['pow ls']
    df_out['pow dz'] = df_merged['pow dz']
    df_out['właściciel'] = df_merged['Właściciel']

    nieotaksowane = xls[~xls['nr_dz'].isin(val['nr_dz'])].copy() if not xls.empty else pd.DataFrame(
        columns=['J. rej.', 'nr_dz', 'Właściciel', 'pow ls', 'pow dz'])
    if not nieotaksowane.empty:
        nieotaksowane = nieotaksowane[['J. rej.', 'nr_dz', 'Właściciel', 'pow ls', 'pow dz']]
        nieotaksowane = nieotaksowane.rename(columns={'Właściciel': 'właściciel'})

    return df_out, nieotaksowane


def wykonaj_makro_vba(df_out, df_braki, tylko_wyrownywanie=False):
    df = df_out.copy()
    df['bg_color'] = ""
    df['font_color'] = ""
    TOLERANCJA = 0.0010

    # 1. WARTOŚCI (bez żadnych kolorów tła)
    # Zmieniamy grupowanie z samego 'nr_dz' na ['nr_dz', 'J. rej.'] aby współwłaściciele nie wpływali na siebie
    for (dz, j_rej), group in df.groupby(['nr_dz', 'J. rej.'], sort=False, dropna=False):
        # Zabezpieczenie przed dublami: sumujemy unikalne kontury, by uniknąć inflacji powierzchni
        unikalne_geo = group.drop_duplicates(subset=['litery'])
        suma_geo = unikalne_geo['pow geo'].sum()

        pow_ewid = group['pow ls'].iloc[0]
        pow_docelowa = group['pow dz'].iloc[0]
        is_new_forest = pd.isna(pow_ewid) or str(pow_ewid).strip() == ""
        df_font = 'FF0000' if is_new_forest else '000000'

        # Checkbox: Wymuszenie wyrównania blokuje tworzenie nadmiarów (zielonych działek)
        if tylko_wyrownywanie:
            nadmiar_sciezka = False
        else:
            nadmiar_sciezka = pd.notna(pow_ewid) and suma_geo > (float(pow_ewid) + 0.1)

        suma_przepisanych = 0.0

        for idx in group.index:
            aktualna_pow = group.at[idx, 'pow geo']
            df.at[idx, 'font_color'] = df_font

            if is_new_forest:
                df.at[idx, 'TU POWSTANĄ DANE'] = aktualna_pow
                continue

            if nadmiar_sciezka:
                if pd.notna(pow_docelowa):
                    reszta = float(pow_docelowa) - suma_przepisanych
                    if reszta > 0:
                        wartosc = min(reszta, aktualna_pow)
                        df.at[idx, 'TU POWSTANĄ DANE'] = round(wartosc, 4)
                        suma_przepisanych += wartosc
                    else:
                        df.at[idx, 'TU POWSTANĄ DANE'] = 0.0000
                else:
                    df.at[idx, 'TU POWSTANĄ DANE'] = aktualna_pow
            else:
                if pd.notna(pow_ewid) and suma_geo != 0:
                    nowa = (aktualna_pow / suma_geo) * float(pow_ewid)
                    zaokr = round(nowa, 4)
                    df.at[idx, 'TU POWSTANĄ DANE'] = zaokr if zaokr != 0 else aktualna_pow
                else:
                    df.at[idx, 'TU POWSTANĄ DANE'] = aktualna_pow

    # 2. DOCIĄGANIE RÓŻNIC ZAOKRĄGLEŃ
    for (dz, j_rej), group in df.groupby(['nr_dz', 'J. rej.'], sort=False, dropna=False):
        pow_ewid = group['pow ls'].iloc[0]
        pow_docelowa = group['pow dz'].iloc[0]

        valid_indices = group[group['TU POWSTANĄ DANE'].notna()].index
        if len(valid_indices) == 0:
            continue
        suma_f = df.loc[valid_indices, 'TU POWSTANĄ DANE'].sum()
        roznica = 0.0
        if pd.notna(pow_docelowa) and str(pow_docelowa).strip() != "":
            pow_j = float(pow_docelowa)
            if suma_f > pow_j:
                roznica = pow_j - suma_f
            elif 0 < (pow_j - suma_f) <= TOLERANCJA:
                roznica = pow_j - suma_f
        if roznica == 0.0 and pd.notna(pow_ewid) and str(pow_ewid).strip() != "":
            pow_i = float(pow_ewid)
            if abs(pow_i - suma_f) > 0 and abs(pow_i - suma_f) <= TOLERANCJA:
                roznica = pow_i - suma_f
        if roznica != 0:
            ostatni_wiersz = valid_indices[-1]
            df.at[ostatni_wiersz, 'TU POWSTANĄ DANE'] = round(
                df.at[ostatni_wiersz, 'TU POWSTANĄ DANE'] + roznica, 4)

    # 3. SZUM -> RÓŻOWY
    rows_to_drop = []
    for idx in df.index:
        val = df.at[idx, 'TU POWSTANĄ DANE']
        pow_ewid = df.at[idx, 'pow ls']
        if pd.notna(val) and val <= 0.004:
            if pd.isna(pow_ewid) or str(pow_ewid).strip() == "":
                rows_to_drop.append(idx)
            else:
                df.at[idx, 'bg_color'] = 'FFB6C1'
    if rows_to_drop:
        df = df.drop(index=rows_to_drop)

    # 4. PRZYBYŁO / UBYŁO
    przybylo_data = []
    ubylo_data = []

    # Checkbox wymusza, by ominąć generowanie arkuszy UBYŁO/PRZYBYŁO i nie mazać na zielono
    if not tylko_wyrownywanie:
        for dz, group in df.groupby('nr_dz', sort=False):
            # Unikalne kontury dla całej działki, aby uniknąć zdublowania sumy
            unikalne_geo = group.drop_duplicates(subset=['litery'])
            suma_f = unikalne_geo['TU POWSTANĄ DANE'].sum() if not unikalne_geo.empty else 0.0

            pow_ewid = group['pow ls'].iloc[0]
            pow_docelowa = group['pow dz'].iloc[0]
            j_rej = group['J. rej.'].iloc[0] if 'J. rej.' in group.columns else ""
            startowy_las = float(pow_ewid) if (pd.notna(pow_ewid) and str(pow_ewid).strip() != "") else 0.0
            roznica = round(suma_f - startowy_las, 4)

            if roznica > 0:
                for idx in group.index:
                    if df.at[idx, 'bg_color'] != 'FFB6C1' and pd.notna(df.at[idx, 'TU POWSTANĄ DANE']):
                        df.at[idx, 'bg_color'] = '00FF00'
                przybylo_data.append({
                    'J. rej.': j_rej, 'nr działki': dz,
                    'aktualna pow ls': round(suma_f, 4), 'ls ewidenca': startowy_las,
                    'ile przybyło': roznica,
                    'pow dz': pow_docelowa if pd.notna(pow_docelowa) else ""
                })
            elif roznica < 0:
                ubylo_data.append({
                    'J. rej.': j_rej, 'nr działki': dz,
                    'aktualna pow ls': round(suma_f, 4), 'ls ewidenca': startowy_las,
                    'ile ubyło': roznica,
                    'pow dz': pow_docelowa if pd.notna(pow_docelowa) else ""
                })

        if not df_braki.empty:
            for _, row in df_braki.iterrows():
                if '[OP]' not in str(row.get('właściciel', '')):
                    pow_ewid = row.get('pow ls', np.nan)
                    pow_doc = row.get('pow dz', np.nan)
                    j_rej = row.get('J. rej.', "")
                    if pd.notna(pow_ewid) and float(pow_ewid) > 0:
                        ubylo_data.append({
                            'J. rej.': j_rej, 'nr działki': row.get('nr_dz', ''),
                            'aktualna pow ls': 0.0, 'ls ewidenca': pow_ewid,
                            'ile ubyło': -float(pow_ewid),
                            'pow dz': pow_doc if pd.notna(pow_doc) else ""
                        })

    return df, pd.DataFrame(przybylo_data), pd.DataFrame(ubylo_data)



def formatuj_arkusz_raportowy(worksheet, tytul, hex_kolor_tytulu):
    worksheet['A1'] = tytul
    worksheet.merge_cells('A1:F1')
    worksheet['A1'].font = Font(size=18, bold=True, color=hex_kolor_tytulu)
    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    thick_bottom = Border(bottom=Side(style='thick', color='000000'))
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))

    max_row = worksheet.max_row
    max_col = 6

    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.border = thick_bottom

    for row in range(3, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if col in [1, 2]:
                cell.alignment = Alignment(horizontal='left')

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in range(2, max_row + 1):
            cell = worksheet.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[col_letter].width = (max_length + 2)

# Struktura WSIE.DBF dokładnie wg specyfikacji MIETEK.EXE (kolejność krytyczna!)
WSIE_FIELDS = [
    ('NAZWA', 'C', 40, 0), ('WOJEW', 'C', 30, 0), ('GMINA', 'C', 30, 0),
    ('STAN_NA', 'D', 8, 0), ('OBOW_OD', 'D', 8, 0), ('OBOW_DO', 'D', 8, 0),
    ('NR_WSI', 'N', 3, 0), ('ROK_ZAL', 'C', 2, 0),
    ('SPR', 'C', 75, 0), ('ZLC', 'C', 40, 0), ('WW', 'C', 20, 0), ('KR', 'C', 5, 0),
    ('DZ1', 'C', 75, 0), ('DZ2', 'C', 75, 0),
    ('ET1', 'N', 6, 0), ('ET2', 'N', 6, 0), ('ET3', 'N', 6, 0), ('ET4', 'N', 6, 0),
    ('ET5', 'N', 6, 0), ('ET6', 'N', 6, 0), ('ET7', 'N', 6, 0),
    ('OCHR2', 'C', 75, 0), ('OCHR3', 'C', 75, 0), ('OCHR4', 'C', 75, 0),
    ('P_OCH', 'N', 12, 4),
    ('ZDR', 'C', 75, 0), ('ZDR1', 'C', 75, 0), ('ZDR2', 'C', 75, 0),
    ('ZG1', 'N', 12, 4), ('ZG2', 'N', 12, 4), ('ZG3', 'N', 12, 4),
    ('PRZY', 'C', 75, 0), ('PRZY1', 'C', 75, 0), ('PRZY2', 'C', 75, 0),
    ('SANITAR', 'C', 75, 0), ('SANITAR1', 'C', 75, 0), ('SANITAR2', 'C', 75, 0),
    ('US1', 'C', 75, 0), ('US2', 'C', 75, 0), ('US3', 'C', 75, 0), ('US4', 'C', 75, 0),
    ('EG1', 'C', 50, 0), ('EG2', 'C', 50, 0), ('EG3', 'C', 50, 0),
    ('EG4', 'C', 50, 0), ('EG5', 'C', 50, 0),
    ('POWIAT', 'C', 30, 0),
]


