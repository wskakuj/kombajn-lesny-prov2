"""
Kombajn Leśny PRO — Mixin: TabRozliczanieMixin
"""

import customtkinter as ctk
from tkinter import messagebox
from pathlib import Path
import threading
import json
import sys
import shutil
import win32com.client

import re
import traceback
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.config import (
    add_tooltip, is_file_locked, load_margins,
)

from app.core.word_worker import (
    run_word_worker,
)

from app.core.excel_tasks import (
    wczytaj_i_przetworz_wlascicieli,
    wczytaj_i_przetworz_val,
    polacz_xls_i_val,
    wykonaj_makro_vba,
    formatuj_arkusz_raportowy,
)

class TabRozliczanieMixin:
    """Mixin dla ModernApp — metody zostały wyciągnięte z oryginalnego guipia.py."""
    pass

    def setup_rozliczanie_tab(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)
        scroll_frame = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        scroll_frame.grid_columnconfigure(0, weight=1)
        font_label = ctk.CTkFont(family="Segoe UI", size=13, weight="bold")
        font_btn = ctk.CTkFont(family="Segoe UI", size=13)
        card = ctk.CTkFrame(
            scroll_frame, fg_color="#252526", corner_radius=8,
            border_width=1, border_color="#333333",
        )
        card.grid(row=0, column=0, padx=20, pady=(15, 15), sticky="new")
        card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(card, text="Folder z plikami XLS:", font=font_label, text_color="#E0E0E0").grid(
            row=0, column=0, padx=15, pady=(15, 8), sticky="w")
        self.rozl_xls_entry = ctk.CTkEntry(
            card, placeholder_text="Wskaż folder z ewidencją (.xls/.xlsx)", height=36)
        self.rozl_xls_entry.grid(row=0, column=1, padx=5, pady=(15, 8), sticky="ew")
        _saved = self.get_setting("folder_rozl_xls_entry")
        if _saved: self.rozl_xls_entry.insert(0, _saved)
        ctk.CTkButton(
            card, text="Przeglądaj", image=self.icon_folder,
            command=lambda: self.select_dir(self.rozl_xls_entry),
            width=110, height=36, font=font_btn, fg_color="#333333", hover_color="#444444",
        ).grid(row=0, column=2, padx=15, pady=(15, 8))

        ctk.CTkLabel(card, text="Folder z plikami VAL:", font=font_label, text_color="#E0E0E0").grid(
            row=1, column=0, padx=15, pady=8, sticky="w")
        self.rozl_val_entry = ctk.CTkEntry(
            card, placeholder_text="Wskaż folder z plikami z geodezji (.val)", height=36)
        self.rozl_val_entry.grid(row=1, column=1, padx=5, pady=8, sticky="ew")
        _saved = self.get_setting("folder_rozl_val_entry")
        if _saved: self.rozl_val_entry.insert(0, _saved)
        ctk.CTkButton(
            card, text="Przeglądaj", image=self.icon_folder,
            command=lambda: self.select_dir(self.rozl_val_entry),
            width=110, height=36, font=font_btn, fg_color="#333333", hover_color="#444444",
        ).grid(row=1, column=2, padx=15, pady=8)

        ctk.CTkLabel(card, text="Folder docelowy:", font=font_label, text_color="#E0E0E0").grid(
            row=2, column=0, padx=15, pady=(8, 15), sticky="w")
        self.rozl_out_entry = ctk.CTkEntry(
            card, placeholder_text="Gdzie zapisać rozliczone tabele?", height=36)
        self.rozl_out_entry.grid(row=2, column=1, padx=5, pady=(8, 15), sticky="ew")
        _saved = self.get_setting("folder_rozl_out_entry")
        if _saved: self.rozl_out_entry.insert(0, _saved)
        ctk.CTkButton(
            card, text="Przeglądaj", image=self.icon_folder,
            command=lambda: self.select_dir(self.rozl_out_entry),
            width=110, height=36, font=font_btn, fg_color="#333333", hover_color="#444444",
        ).grid(row=2, column=2, padx=15, pady=(8, 15))

        ctk.CTkLabel(
            card,
            text="Dopasowanie: nazwa pliku XLS → plik *.val o tej samej nazwie (ignorując spacje i _ )",
            font=ctk.CTkFont(family="Segoe UI", size=12), text_color="#888888",
        ).grid(row=3, column=0, columnspan=3, padx=15, pady=(0, 5), sticky="w")

        self.rozl_tylko_wyrownywanie_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            card, text="Wymuś proporcjonalne wyrównanie do ewidencji Ls (nie twórz arkuszy PRZYBYŁO/UBYŁO)",
            variable=self.rozl_tylko_wyrownywanie_var, font=ctk.CTkFont(family="Segoe UI", size=12),
            fg_color="#0067C0", hover_color="#005A9E"
        ).grid(row=4, column=0, columnspan=3, padx=15, pady=(0, 5), sticky="w")

        # --- NOWY CHECKBOX ---
        self.rozl_usun_puste_jrej_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            card, text="Usuń wiersze jeśli brakuje wartości w J. rej. (usuwanie wydzieleń bez właścicieli)",
            variable=self.rozl_usun_puste_jrej_var, font=ctk.CTkFont(family="Segoe UI", size=12),
            fg_color="#8B0000", hover_color="#A52A2A"
        ).grid(row=5, column=0, columnspan=3, padx=15, pady=(0, 15), sticky="w")
        # ---------------------

        self.rozl_start_btn = ctk.CTkButton(
            scroll_frame, text="Uruchom rozliczanie obrębów", image=self.icon_start,
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            fg_color="#0067C0", hover_color="#005A9E", height=44, corner_radius=6,
            command=self.start_rozliczanie_pipeline,
        )
        self.rozl_start_btn.grid(row=1, column=0, padx=20, pady=(5, 20), sticky="ew")
        add_tooltip(
            self.rozl_start_btn,
            "Łączy ewidencję XLS z plikami VAL geodezji, przelicza powierzchnie "
            "i zapisuje raporty z arkuszami: Tabela_Glowna, Nieotaksowane, PRZYBYLO, UBYLO.",
        )

        # ==========================================
        # LEGENDA KOLORÓW  (Tabela_Glowna, kolumna F)
        # ==========================================
        legend_card = ctk.CTkFrame(
            scroll_frame, fg_color="#252526", corner_radius=8,
            border_width=1, border_color="#333333",
        )
        legend_card.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="new")
        legend_card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            legend_card, text="Legenda kolorów",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            text_color="#E0E0E0",
        ).grid(row=0, column=0, padx=15, pady=(14, 2), sticky="w")
        ctk.CTkLabel(
            legend_card,
            text="Kolory opisują, w jaki sposób została wyliczona powierzchnia w danej komórce. Najedź kursorem na wiersz.",
            font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#888888",
        ).grid(row=1, column=0, padx=15, pady=(0, 10), sticky="w")

        # --- dane legendy: (hex, nazwa, opis, rodzaj) ; rodzaj = "fill" albo "font" ---
        fill_rows = [
            ("00FF00", "ZIELONY",
             "Działka „przybyła” (suma w kolumnie F > ewidencja LS z kolumny I) — podświetlone wszystkie jej wiersze z wartością."),
            ("FFB6C1", "RÓŻOWY",
             "Pomijalny „szum” (≤ 0,004 ha) przy istniejącej ewidencji LS — do sprawdzenia, czy wywalić."),
            ("FFFFFF", "BRAK WYPEŁNIENIA",
             "Wiersz rozliczony standardowo (działka ani nie przybyła, ani nie jest szumem)."),
        ]
        font_rows = [
            ("FF0000", "CZERWONY TEKST",
             "W kolumnie I (ewidencja LS) NIE MA wartości — liczba w F pochodzi tylko z geomapy (sprawdź ręcznie)."),
        ]

        ROW_BG, ROW_HOVER = "#202022", "#2E2E31"

        def _legend_row(parent, r, hex_color, title, desc, kind):
            row_frame = ctk.CTkFrame(
                parent, fg_color=ROW_BG, corner_radius=6,
                border_width=1, border_color="#2C2C2E",
            )
            row_frame.grid(row=r, column=0, padx=12, pady=3, sticky="ew")
            row_frame.grid_columnconfigure(2, weight=1)

            if kind == "fill":
                swatch = ctk.CTkFrame(
                    row_frame, width=30, height=18, corner_radius=3,
                    fg_color=f"#{hex_color}", border_width=1, border_color="#5A5A5A",
                )
                swatch.grid(row=0, column=0, padx=(10, 10), pady=8, sticky="w")
                swatch.grid_propagate(False)
            else:
                swatch = ctk.CTkLabel(
                    row_frame, text="Aa", width=30,
                    font=ctk.CTkFont(family="Consolas", size=13, weight="bold"),
                    text_color=f"#{hex_color}", fg_color="transparent",
                )
                swatch.grid(row=0, column=0, padx=(10, 10), pady=8, sticky="w")

            ctk.CTkLabel(
                row_frame, text=title,
                font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
                text_color="#F0F0F0", width=150, anchor="w", fg_color="transparent",
            ).grid(row=0, column=1, padx=(0, 8), pady=8, sticky="w")
            ctk.CTkLabel(
                row_frame, text=desc,
                font=ctk.CTkFont(family="Segoe UI", size=11),
                text_color="#B8B8B8", anchor="w", wraplength=520, justify="left",
                fg_color="transparent",
            ).grid(row=0, column=2, padx=(0, 10), pady=8, sticky="w")

            # mikrointerakcja: cały wiersz rozjaśnia się pod kursorem
            def _bind_hover(widget):
                widget.bind("<Enter>", lambda e: row_frame.configure(fg_color=ROW_HOVER))
                widget.bind("<Leave>", lambda e: row_frame.configure(fg_color=ROW_BG))
                for child in widget.winfo_children():
                    _bind_hover(child)

            _bind_hover(row_frame)

        # --- nagłówek + wiersze sekcji WYPEŁNIENIE ---
        ctk.CTkLabel(
            legend_card, text="▮  WYPEŁNIENIE KOMÓRKI (tło)",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color="#0078D7",
        ).grid(row=2, column=0, padx=15, pady=(4, 2), sticky="w")
        row_idx = 3
        for hex_color, title, desc in fill_rows:
            _legend_row(legend_card, row_idx, hex_color, title, desc, "fill")
            row_idx += 1

        # --- nagłówek + wiersze sekcji KOLOR TEKSTU ---
        ctk.CTkLabel(
            legend_card, text="A  KOLOR TEKSTU (czcionka)",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color="#E0A020",
        ).grid(row=row_idx, column=0, padx=15, pady=(10, 2), sticky="w")
        row_idx += 1
        for hex_color, title, desc in font_rows:
            _legend_row(legend_card, row_idx, hex_color, title, desc, "font")
            row_idx += 1

        ctk.CTkLabel(
            legend_card,
            text="Podpowiedź: czerwone i szare komórki warto przejrzeć ręcznie przed wstrzyknięciem krzyżówek.",
            font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#888888",
        ).grid(row=row_idx, column=0, padx=15, pady=(10, 14), sticky="w")

    def validate_rozliczanie(self, folder_xls, folder_val, folder_out):
        """Walidacja przed startem — sprawdza foldery, pliki i dopasowanie. Zwraca listę problemów."""
        problems = []

        # 1. Sprawdź foldery
        if not folder_xls:
            problems.append("• Nie wskazano folderu XLS.")
        elif not Path(folder_xls).exists():
            problems.append(f"• Folder XLS nie istnieje:\n  {folder_xls}")

        if not folder_val:
            problems.append("• Nie wskazano folderu VAL.")
        elif not Path(folder_val).exists():
            problems.append(f"• Folder VAL nie istnieje:\n  {folder_val}")

        if not folder_out:
            problems.append("• Nie wskazano folderu wyjściowego.")

        if problems:
            return problems  # Nie sprawdzaj dalej jeśli foldery są złe

        # 2. Sprawdź pliki XLS
        xls_files = sorted([p for p in Path(folder_xls).iterdir()
                     if p.is_file() and p.suffix.lower() in {".xls", ".xlsx"} and not p.name.startswith("~$")])
        if not xls_files:
            problems.append("• Folder XLS nie zawiera żadnych plików .xls/.xlsx.")
            return problems

        # 3. Sprawdź pliki VAL
        val_files = sorted([p for p in Path(folder_val).iterdir() if p.is_file() and p.suffix.lower() == ".val"])
        if not val_files:
            problems.append("• Folder VAL nie zawiera żadnych plików .val.")
            return problems

        # 4. Dopasowanie XLS ↔ VAL — używa tej samej logiki co run_rozliczanie_thread
        #    czyli: wyczyść nazwę ze spacji/podkreślników i sprawdź czy nazwa pliku VAL
        #    kończy się na nazwę XLS + ".val" (np. 0301BIAŁCZ.val pasuje do BIAŁCZ.xls)
        import re as _re
        val_matched = set()
        xls_without_val = []
        for x in xls_files:
            x_czysta = _re.sub(r"[\s_]", "", x.stem.lower())
            found = False
            for v in val_files:
                v_czysta = _re.sub(r"[\s_]", "", v.name.lower())
                if v_czysta.endswith(x_czysta + ".val"):
                    found = True
                    val_matched.add(v.name)
                    break
            if not found:
                xls_without_val.append(x.stem)

        val_without_xls = [v.stem for v in val_files if v.name not in val_matched]

        if xls_without_val:
            xls_without_val.sort()
            problems.append(f"• Brak plików .val dla {len(xls_without_val)} obrębów:")
            for nazwa in xls_without_val[:10]:
                problems.append(f"    — {nazwa} (brak pasującego .val)")
            if len(xls_without_val) > 10:
                problems.append(f"    ... i {len(xls_without_val) - 10} więcej")

        if val_without_xls:
            val_without_xls.sort()
            problems.append(f"• Pliki .val bez odpowiadających .xls ({len(val_without_xls)}):")
            for nazwa in val_without_xls[:10]:
                problems.append(f"    — {nazwa}.val (brak pasującego .xls)")
            if len(val_without_xls) > 10:
                problems.append(f"    ... i {len(val_without_xls) - 10} więcej")

        return problems

    def start_rozliczanie_pipeline(self):
        folder_xls = self.rozl_xls_entry.get().strip() if self.rozl_xls_entry else ""
        folder_val = self.rozl_val_entry.get().strip() if self.rozl_val_entry else ""
        folder_out = self.rozl_out_entry.get().strip() if self.rozl_out_entry else ""

        if not folder_xls or not Path(folder_xls).exists():
            messagebox.showwarning("Błąd", "Wybierz istniejący folder z plikami XLS.")
            return
        if not folder_val or not Path(folder_val).exists():
            messagebox.showwarning("Błąd", "Wybierz istniejący folder z plikami VAL.")
            return
        if not folder_out:
            messagebox.showwarning("Błąd", "Wybierz folder docelowy dla raportów.")
            return
        if self.running:
            return

        tylko_wyrownywanie = self.rozl_tylko_wyrownywanie_var.get()

        tylko_wyrownywanie = self.rozl_tylko_wyrownywanie_var.get()
        usun_puste_jrej = self.rozl_usun_puste_jrej_var.get()

        # Walidacja przed startem
        problems = self.validate_rozliczanie(folder_xls, folder_val, folder_out)
        if problems:
            msg = "Wykryto problemy przed startem:\n\n" + "\n".join(problems)
            msg += "\n\nCzy chcesz kontynuować mimo to?"
            if not messagebox.askyesno("Walidacja — znaleziono problemy", msg):
                return

        # Zapisz foldery w ustawieniach
        self.set_setting("folder_rozl_xls_entry", folder_xls)
        self.set_setting("folder_rozl_val_entry", folder_val)
        self.set_setting("folder_rozl_out_entry", folder_out)

        self.last_output_dir = Path(folder_out)
        self._disable_ui_for_process()
        self.log(f"[ROZLICZANIE] URUCHOMIENIE PROCEDURY\nXLS: {folder_xls}\nVAL: {folder_val}")
        self.set_progress(0)
        threading.Thread(
            target=self.run_rozliczanie_thread,
            args=(folder_xls, folder_val, folder_out, tylko_wyrownywanie, usun_puste_jrej),
            daemon=True,
        ).start()


    def run_rozliczanie_thread(self, folder_xls_str, folder_val_str, folder_out_str, tylko_wyrownywanie, usun_puste_jrej=False):
        try:
            folder_xls = Path(folder_xls_str)
            folder_val = Path(folder_val_str)
            folder_out = Path(folder_out_str)
            folder_out.mkdir(parents=True, exist_ok=True)

            xls_files = sorted(
                [
                    p for p in folder_xls.iterdir()
                    if p.is_file()
                       and p.suffix.lower() in {".xls", ".xlsx"}
                       and not p.name.startswith("~$")
                ]
            )
            if not xls_files:
                raise Exception("Brak plików XLS/XLSX we wskazanym folderze.")

            total = len(xls_files)
            self.start_progress_tracking(total, "Rozliczanie obrębów")
            self.update_status("Rozliczanie powierzchni obrębów...", "#0078D7")

            stat_sukces = 0
            stat_brak_val = []
            stat_bledy = []

            for idx, sciezka_xls in enumerate(xls_files, start=1):
                self.check_stop()
                nazwa_wsi = sciezka_xls.stem
                self.log(f"[ROZLICZANIE] ▸ Przetwarzanie obrębu: {nazwa_wsi}")

                # PRECYZYJNE DOPASOWANIE VAL (zapobiega pomyleniu "LIS" z "LISIE_POLE")
                nazwa_wsi_czysta = re.sub(r"[\s_]", "", nazwa_wsi.lower())
                pasujace_val = []
                for f in folder_val.iterdir():
                    if f.is_file() and f.suffix.lower() == ".val":
                        f_czysta = re.sub(r"[\s_]", "", f.name.lower())
                        if f_czysta.endswith(nazwa_wsi_czysta + ".val"):
                            pasujace_val.append(f)

                if not pasujace_val:
                    self.log(f"  ⚠️ Pominięto '{nazwa_wsi}' — brak pasującego pliku .val")
                    stat_brak_val.append(nazwa_wsi)
                    self.set_progress(idx / total, current_file=sciezka_xls.name, current=idx)
                    continue

                sciezka_val = pasujace_val[0]
                plik_wyjsciowy = folder_out / f"{nazwa_wsi}_Rozliczone.xlsx"

                try:
                    tabela_xls, df_full = wczytaj_i_przetworz_wlascicieli(str(sciezka_xls))
                    tabela_val = wczytaj_i_przetworz_val(str(sciezka_val))

                    if tabela_val is None:
                        raise Exception(f"Nie udało się wczytać pliku VAL: {sciezka_val.name}")

                    tabela_glowna, tabela_braki = polacz_xls_i_val(tabela_xls, df_full, tabela_val)

                    # LOGIKA USUWANIA WIERSZY BEZ J. REJ.
                    if usun_puste_jrej:
                        # Rzutujemy na liczby, traktując wszelkie braki jako '0'
                        jrej_num = pd.to_numeric(tabela_glowna['J. rej.'], errors='coerce').fillna(0)
                        # Nadpisujemy tabelę pozostawiając tylko wiersze, gdzie J. rej. NIE JEST ZEREM
                        tabela_glowna = tabela_glowna[jrej_num != 0].copy()

                    tabela_gotowa, tabela_przybylo, tabela_ubylo = wykonaj_makro_vba(
                        tabela_glowna, tabela_braki, tylko_wyrownywanie=tylko_wyrownywanie)

                    with pd.ExcelWriter(str(plik_wyjsciowy), engine="openpyxl") as writer:
                        kolumny_wyjsciowe = [
                            c for c in tabela_gotowa.columns
                            if c not in ("bg_color", "font_color")
                        ]
                        tabela_gotowa[kolumny_wyjsciowe].to_excel(
                            writer, sheet_name="Tabela_Glowna", index=False)

                        if not tabela_braki.empty:
                            tabela_braki_eksport = tabela_braki[
                                ["J. rej.", "nr_dz", "pow ls", "pow dz", "właściciel"]]
                            tabela_braki_eksport.to_excel(
                                writer, sheet_name="Nieotaksowane", index=False)
                        else:
                            pd.DataFrame(
                                columns=["J. rej.", "nr_dz", "pow ls", "pow dz", "właściciel"]
                            ).to_excel(writer, sheet_name="Nieotaksowane", index=False)

                        if not tabela_przybylo.empty:
                            tabela_przybylo.to_excel(
                                writer, sheet_name="PRZYBYLO", index=False, startrow=1)
                        if not tabela_ubylo.empty:
                            tabela_ubylo.to_excel(
                                writer, sheet_name="UBYLO", index=False, startrow=1)

                        # Kolorowanie kolumny "TU POWSTANĄ DANE" (kolumna 6)
                        worksheet_glowna = writer.sheets["Tabela_Glowna"]
                        for row_idx, row in enumerate(tabela_gotowa.itertuples(), start=2):
                            bg_col = getattr(row, "bg_color", "")
                            f_col = getattr(row, "font_color", "")
                            cell = worksheet_glowna.cell(row=row_idx, column=6)
                            if pd.notna(bg_col) and bg_col != "":
                                cell.fill = PatternFill(
                                    start_color=str(bg_col), end_color=str(bg_col),
                                    fill_type="solid")
                            if pd.notna(f_col) and f_col != "":
                                cell.font = Font(color=str(f_col))

                        if "PRZYBYLO" in writer.sheets:
                            formatuj_arkusz_raportowy(
                                writer.sheets["PRZYBYLO"], "PRZYBYŁO", "FF0000")
                        if "UBYLO" in writer.sheets:
                            formatuj_arkusz_raportowy(
                                writer.sheets["UBYLO"], "UBYŁO", "87CEEB")

                    self.log(f"  ✅ Zapisano: {plik_wyjsciowy.name}")
                    stat_sukces += 1

                except PermissionError:
                    self.log(
                        f"  ❌ BŁĄD UPRAWNIEŃ: nie można zapisać '{nazwa_wsi}_Rozliczone.xlsx' "
                        f"— zamknij plik w Excelu i spróbuj ponownie.")
                    stat_bledy.append(nazwa_wsi)
                except Exception as e:
                    self.log(f"  ❌ Błąd przy obrębie '{nazwa_wsi}': {e}")
                    stat_bledy.append(nazwa_wsi)

                self.set_progress(idx / total, current_file=sciezka_xls.name, current=idx)

            # --- RAPORT KOŃCOWY ---
            self.log("\n" + "=" * 55)
            self.log("PODSUMOWANIE ROZLICZANIA OBRĘBÓW")
            self.log(f"✅ Przetworzono poprawnie: {stat_sukces}")
            if stat_brak_val:
                self.log(f"⚠️ Pominięto (brak VAL): {len(stat_brak_val)} -> {', '.join(stat_brak_val)}")
            if stat_bledy:
                self.log(f"❌ Zakończone błędem: {len(stat_bledy)} -> {', '.join(stat_bledy)}")
            self.log("=" * 55)

            self.update_status("Zakończono pomyślnie.", "#27ae60", animate=False)
            podsumowanie = (
                f"Rozliczanie obrębów zakończone.\n\n"
                f"✅ Przetworzono poprawnie: {stat_sukces}\n"
                f"⚠️ Pominięto (brak pliku VAL): {len(stat_brak_val)}\n"
                f"❌ Zakończone błędem: {len(stat_bledy)}"
            )
            self.after(0, lambda: messagebox.showinfo("Rozliczanie obrębów", podsumowanie))

        except InterruptedError:
            self.update_status("Przerwano", "#D83B01", animate=False)
            self.log("\nZADANIE PRZERWANE PRZEZ UŻYTKOWNIKA.")
        except Exception as e:
            self.log(traceback.format_exc())
            self.update_status("Błąd", "#D83B01", animate=False)
        finally:
            self.running = False
            self.after(0, self.restore_all_buttons)

    def _build_margins_ui(self, parent_frame, row_idx, mode_key):
        if not hasattr(self, "margin_vars"):
            self.margin_vars = {}
        self.margin_vars[mode_key] = {}

        font_label = ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
        font_entry = ctk.CTkFont(family="Segoe UI", size=11)

        margin_frame = ctk.CTkFrame(parent_frame, fg_color="#1E1E1E", border_width=1, border_color="#333333")
        margin_frame.grid(row=row_idx, column=0, columnspan=3, padx=15, pady=(5, 15), sticky="ew")

        ctk.CTkLabel(margin_frame, text="Ustawienia marginesów (w cm):", font=font_label, text_color="#A0A0A0").grid(
            row=0, column=0, columnspan=5, pady=(5, 5), sticky="w", padx=10)

        headers = ["Typ pliku", "Góra", "Dół", "Lewo", "Prawo"]
        for c, h in enumerate(headers):
            ctk.CTkLabel(margin_frame, text=h, font=font_label, text_color="#0078D7").grid(row=1, column=c, padx=5,
                                                                                           pady=(0, 5))

        file_types = ["REJESTR1", "OPTAX", "TAB_KLW3", "WSKAZ1", "HALIZNY", "WYK_NEG", "OPIS", "ZEST1", "WK_ZM1"]

        # --- ZMIANA: Pobranie zapisanych marginesów z pliku ---
        saved_config = load_margins()
        mode_saved = saved_config.get(mode_key, {})

        for r, ftype in enumerate(file_types, start=2):
            ctk.CTkLabel(margin_frame, text=ftype, font=font_entry).grid(row=r, column=0, padx=10, pady=2, sticky="w")

            # Pobieramy konkretne wartości dla pliku, jeśli brak to stosujemy twarde domyślne
            file_saved = mode_saved.get(ftype, {})
            t_val = str(file_saved.get("T", "1.5"))
            b_val = str(file_saved.get("B", "1.5"))
            l_val = str(file_saved.get("L", "2.5"))
            r_val = str(file_saved.get("R", "1.5"))

            eT = ctk.CTkEntry(margin_frame, width=45, height=24, font=font_entry)
            eT.insert(0, t_val)
            eT.grid(row=r, column=1, padx=5, pady=2)

            eB = ctk.CTkEntry(margin_frame, width=45, height=24, font=font_entry)
            eB.insert(0, b_val)
            eB.grid(row=r, column=2, padx=5, pady=2)

            eL = ctk.CTkEntry(margin_frame, width=45, height=24, font=font_entry)
            eL.insert(0, l_val)
            eL.grid(row=r, column=3, padx=5, pady=2)

            eR = ctk.CTkEntry(margin_frame, width=45, height=24, font=font_entry)
            eR.insert(0, r_val)
            eR.grid(row=r, column=4, padx=5, pady=2)

            self.margin_vars[mode_key][ftype] = {"T": eT, "B": eB, "L": eL, "R": eR}

    # NOWA METODA: Konfiguracja UI dla Pełny Automat (STR_TYT + SKROTY)
    def start_remove_columns_pipeline(self):
        folder = self.excel_folder_entry.get().strip() if self.excel_folder_entry else ""
        output_folder = self.excel_output_entry.get().strip() if self.excel_output_entry else ""

        if not folder or not Path(folder).exists():
            messagebox.showwarning("Błąd", "Wybierz istniejący folder źródłowy z plikami Excel.")
            return
        if not output_folder:
            messagebox.showwarning("Błąd", "Wybierz folder docelowy dla zapisanych plików.")
            return

        remove_owners = self.remove_owners_var.get()
        remove_ls = self.remove_ls_var.get()

        if not remove_owners and not remove_ls:
            messagebox.showwarning("Brak wyboru", "Zaznacz przynajmniej jedną opcję usuwania (Właściciele lub LS).")
            return

        if self.running:
            return

        self.last_output_dir = Path(output_folder)
        self._disable_ui_for_process()
        self.log(f"[EXCEL] URUCHOMIENIE: usuwanie określonych kolumn z arkuszy Sheet4 / REJ\nZ: {folder}")
        self.set_progress(0)

        include_subfolders = (
                getattr(self, "include_subfolders_var", None)
                and self.include_subfolders_var.get()
        )
        threading.Thread(
            target=self.run_remove_columns_thread,
            args=(folder, output_folder, include_subfolders, remove_owners, remove_ls),
            daemon=True,
        ).start()

    def run_remove_columns_thread(self, folder_str, output_folder_str, include_subfolders, remove_owners, remove_ls):
        import pythoncom
        pythoncom.CoInitialize()
        excel = None
        try:
            folder = Path(folder_str)
            output_folder = Path(output_folder_str)

            files = (
                list(folder.rglob("*.xls*"))
                if include_subfolders
                else list(folder.glob("*.xls*"))
            )
            files = sorted([f for f in files if f.is_file() and not f.name.startswith("~$")])

            if not files:
                raise Exception("Brak plików Excel.")

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            total = len(files)

            # Lista wartości do usunięcia
            values_to_delete = []
            if remove_owners: values_to_delete.append(2)
            if remove_ls: values_to_delete.append(3)

            self.start_progress_tracking(total, "Usuwanie kolumn")

            for idx, file_path in enumerate(files, start=1):
                self.check_stop()
                self.progress_current_file = file_path.name
                if is_file_locked(file_path):
                    self.log(f"POMINIĘTO (Plik zablokowany/otwarty): {file_path.name}")
                    continue

                self.log(f"Przetwarzanie (usuwanie kolumn): {file_path.name}")
                wb = None
                try:
                    rel_path = file_path.relative_to(folder)
                    target_path = output_folder / rel_path
                    target_path.parent.mkdir(parents=True, exist_ok=True)

                    if file_path.resolve() != target_path.resolve():
                        shutil.copy2(file_path, target_path)

                    wb = excel.Workbooks.Open(str(target_path))
                    wb.CheckCompatibility = False

                    # --- LOGIKA USUWANIA KOLUMN ---
                    for sheet_name in ["Sheet4", "REJ"]:
                        try:
                            ws = self.get_sheet_if_exists(wb, sheet_name)
                            if ws:
                                # Skrypt sprawdza kolumny od 50 w dół, do 1.
                                # Robimy to od tyłu, żeby przesunięcie kolumn (po usunięciu)
                                # nie popsuło indeksów dla pozostałych kolumn.
                                for col in range(50, 0, -1):
                                    cell_val = ws.Cells(9, col).Value
                                    if cell_val is not None:
                                        try:
                                            # Rzutujemy ew. wartość float 2.0 na integer 2
                                            val_int = int(float(cell_val))
                                            if val_int in values_to_delete:
                                                ws.Columns(col).Delete()
                                                self.log(
                                                    f"  -> Usunięto kolumnę {col} (znaleziono wartość {val_int} w wierszu 9) z arkusza {sheet_name}")
                                        except Exception:
                                            pass
                        except Exception as e:
                            self.log(f"  [Ostrzeżenie] Problem z usunięciem w {sheet_name}: {e}")

                    wb.Close(SaveChanges=True)
                except Exception as e:
                    self.log(f"Błąd pliku {file_path.name}: {e}")

                if wb is not None:
                    try:
                        wb.Close(SaveChanges=False)
                    except:
                        pass
                self.set_progress(idx / total)

            self.update_status("Zakończono pomyślnie.", "#27ae60", animate=False)
        except InterruptedError:
            self.update_status("Przerwano", "#D83B01", animate=False)
            self.log("\nZADANIE PRZERWANE PRZEZ UŻYTKOWNIKA.")
        except Exception as e:
            self.log(traceback.format_exc())
            self.update_status("Błąd", "#D83B01", animate=False)
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except:
                    pass
            pythoncom.CoUninitialize()
            self.running = False
            self.after(0, self.restore_all_buttons)


if __name__ == "__main__":
    if "--word-worker" in sys.argv:
        log_file_path = None
        if "--log-file" in sys.argv:
            l_idx = sys.argv.index("--log-file")
            log_file_path = sys.argv[l_idx + 1]


            class FileLogger:
                def __init__(self, filename):
                    self.filename = filename
                    with open(self.filename, "w", encoding="utf-8") as f:
                        f.write("")

                def write(self, text):
                    with open(self.filename, "a", encoding="utf-8") as f:
                        f.write(str(text))

                def flush(self):
                    pass


            sys.stdout = sys.stderr = FileLogger(log_file_path)

        try:
            idx = sys.argv.index("--word-worker")
            in_dir = sys.argv[idx + 1]
            out_dir = sys.argv[idx + 2]
            remove_names = "--remove-names" in sys.argv

            file_filter = []
            idx_scan = 0
            margins_config = {}
            while idx_scan < len(sys.argv):
                if sys.argv[idx_scan] == "--filter" and idx_scan + 1 < len(sys.argv):
                    file_filter.append(sys.argv[idx_scan + 1])
                    idx_scan += 2
                    continue
                if sys.argv[idx_scan] == "--margins-file" and idx_scan + 1 < len(sys.argv):
                    try:
                        with open(sys.argv[idx_scan + 1], 'r', encoding='utf-8') as mf:
                            margins_config = json.load(mf)
                    except:
                        pass
                    idx_scan += 2
                    continue
                idx_scan += 1

            if not file_filter:
                file_filter = ["Wszystkie"]

            run_word_worker(in_dir, out_dir, remove_names, file_filter, margins_config)

        except Exception as e:
            # Wychwytujemy WSZYSTKIE błędy procesu w tle i zapisujemy je do logu!
            import traceback

            if log_file_path:
                try:
                    with open(log_file_path, "a", encoding="utf-8") as f:
                        f.write(f"\n[BŁĄD KRYTYCZNY PROCESU WORD]: {e}\n")
                        f.write(traceback.format_exc())
                except:
                    pass
            sys.exit(1)

        sys.exit(0)

    app = ModernApp()
    app.mainloop()
