"""
Kombajn Leśny PRO — Mixin: TabMietekRozbieznosciMixin
"""

import customtkinter as ctk
from tkinter import messagebox
from pathlib import Path
import threading
import re
import warnings
warnings.filterwarnings("ignore", message=".*OLE2 inconsistency.*")
warnings.filterwarnings("ignore", message=".*file size.*not.*sector size.*")
warnings.filterwarnings("ignore", message=".*SSCS size.*")
import traceback
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Side

from app.config import (
    add_tooltip,
)

from app.core.word_worker import (
    get_resource_path,
)

from app.core.excel_tasks import (
    wczytaj_i_przetworz_wlascicieli,
)

class TabMietekRozbieznosciMixin:
    """Mixin dla ModernApp — metody zostały wyciągnięte z oryginalnego guipia.py."""
    pass

    def setup_mietek_rozbieznosci_tab(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)
        scroll_frame = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        scroll_frame.grid_columnconfigure(0, weight=1)
        font_label = ctk.CTkFont(family="Segoe UI", size=13, weight="bold")
        font_btn = ctk.CTkFont(family="Segoe UI", size=13)
        card = ctk.CTkFrame(
            scroll_frame, fg_color="#252526", corner_radius=8,
            border_width=1, border_color="#333333",
        )
        card.grid(row=0, column=0, padx=20, pady=(15, 15), sticky="new")
        card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            card, text="1. Folder z Mietkami (D*.DBF):", font=font_label, text_color="#E0E0E0",
        ).grid(row=0, column=0, padx=15, pady=(15, 8), sticky="w")
        self.mietek_rozb_mietki_entry = ctk.CTkEntry(
            card, placeholder_text="Wskaż folder zawierający pliki D*.DBF (nawet bezpośrednio)", height=36,
        )
        self.mietek_rozb_mietki_entry.grid(row=0, column=1, padx=5, pady=(15, 8), sticky="ew")
        ctk.CTkButton(
            card, text="Przeglądaj", image=self.icon_folder,
            command=lambda: self.select_dir(self.mietek_rozb_mietki_entry),
            width=110, height=36, font=font_btn, fg_color="#333333", hover_color="#444444",
        ).grid(row=0, column=2, padx=15, pady=(15, 8))

        ctk.CTkLabel(
            card, text="2. Folder z plikami XLS (Ewidencja):", font=font_label, text_color="#E0E0E0",
        ).grid(row=1, column=0, padx=15, pady=(8, 8), sticky="w")
        self.mietek_rozb_excel_entry = ctk.CTkEntry(
            card, placeholder_text="Wskaż folder z ewidencją (.xls/.xlsx)", height=36,
        )
        self.mietek_rozb_excel_entry.grid(row=1, column=1, padx=5, pady=(8, 8), sticky="ew")
        ctk.CTkButton(
            card, text="Przeglądaj", image=self.icon_folder,
            command=lambda: self.select_dir(self.mietek_rozb_excel_entry),
            width=110, height=36, font=font_btn, fg_color="#333333", hover_color="#444444",
        ).grid(row=1, column=2, padx=15, pady=(8, 8))

        ctk.CTkLabel(
            card, text="3. Folder docelowy zapisu:", font=font_label, text_color="#E0E0E0",
        ).grid(row=2, column=0, padx=15, pady=(8, 8), sticky="w")
        self.mietek_rozb_out_entry = ctk.CTkEntry(
            card, placeholder_text="Gdzie zapisać gotowe raporty Wykazu Rozbieżności?", height=36,
        )
        self.mietek_rozb_out_entry.grid(row=2, column=1, padx=5, pady=(8, 8), sticky="ew")
        ctk.CTkButton(
            card, text="Przeglądaj", image=self.icon_folder,
            command=lambda: self.select_dir(self.mietek_rozb_out_entry),
            width=110, height=36, font=font_btn, fg_color="#333333", hover_color="#444444",
        ).grid(row=2, column=2, padx=15, pady=(8, 8))

        ctk.CTkLabel(
            card,
            text="Powierzchnie w plikach Excel są odczytywane jako m² i przeliczane na ha (÷10000). "
                 "Nazwa obrębu pobierana z WSIE.DBF (pole NAZWA). Tabela sortowana po J. rej. rosnąco.",
            font=ctk.CTkFont(family="Segoe UI", size=12), text_color="#888888",
        ).grid(row=3, column=0, columnspan=3, padx=15, pady=(0, 15), sticky="w")

        # --- DWA PRZYCISKI OBOK SIEBIE ---
        btn_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        btn_frame.grid(row=1, column=0, padx=20, pady=(5, 20), sticky="ew")
        btn_frame.grid_columnconfigure((0, 1), weight=1)

        self.mietek_rozb_start_btn = ctk.CTkButton(
            btn_frame, text="Generuj Wykaz Rozbieżności", image=self.icon_start,
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            fg_color="#0067C0", hover_color="#005A9E", height=44, corner_radius=6,
            command=lambda: self.start_mietek_rozbieznosci_pipeline(bez_nazwisk=False),
        )
        self.mietek_rozb_start_btn.grid(row=0, column=0, padx=(0, 5), sticky="ew")

        self.mietek_rozb_bez_nazwisk_btn = ctk.CTkButton(
            btn_frame, text="Bez Nazwisk", image=self.icon_start,
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            fg_color="#8B0000", hover_color="#A52A2A", height=44, corner_radius=6,
            command=lambda: self.start_mietek_rozbieznosci_pipeline(bez_nazwisk=True),
        )
        self.mietek_rozb_bez_nazwisk_btn.grid(row=0, column=1, padx=(5, 0), sticky="ew")
        add_tooltip(
            self.mietek_rozb_bez_nazwisk_btn,
            "Generuje identyczną tabelę, ale z pustą kolumną właściciela. "
            "Plik zapisuje się z dopiskiem '_BEZ NAZWISK', więc nie nadpisze wersji pełnej.",
        )

    def start_mietek_rozbieznosci_pipeline(self, bez_nazwisk=False):
        mietki_dir = self.mietek_rozb_mietki_entry.get().strip() if self.mietek_rozb_mietki_entry else ""
        excel_dir = self.mietek_rozb_excel_entry.get().strip() if self.mietek_rozb_excel_entry else ""
        out_dir = self.mietek_rozb_out_entry.get().strip() if self.mietek_rozb_out_entry else ""

        if not mietki_dir or not Path(mietki_dir).exists():
            messagebox.showwarning("Błąd", "Wybierz istniejący folder z Mietkami (D*.DBF).")
            return
        if not excel_dir or not Path(excel_dir).exists():
            messagebox.showwarning("Błąd", "Wybierz istniejący folder z plikami XLS Ewidencji.")
            return
        if not out_dir:
            messagebox.showwarning("Błąd", "Wybierz folder docelowy dla raportów.")
            return
        if self.running:
            return

        self.last_output_dir = Path(out_dir)
        self._disable_ui_for_process()
        tryb = "BEZ NAZWISK" if bez_nazwisk else "PEŁNY"
        self.log(f"[WYKAZ ROZBIEŻNOŚCI - {tryb}] URUCHOMIENIE\nMIETKI: {mietki_dir}\nEXCEL: {excel_dir}")
        self.set_progress(0)
        threading.Thread(
            target=self.run_mietek_rozbieznosci_thread,
            args=(mietki_dir, excel_dir, out_dir, bez_nazwisk),
            daemon=True,
        ).start()

    def read_wsie_nazwa(self, dbf_path):
        """Zwraca nazwę obrębu z WSIE.DBF (pole NAZWA pierwszego rekordu,
        czyli odpowiednik 'komórki A2'). Szuka WSIE.DBF obok pliku D*.DBF
        oraz rekurencyjnie w folderze obrębu. Zwraca None, gdy nie znaleziono."""
        try:
            start_dir = Path(dbf_path).parent
            root = start_dir.parent if start_dir.name.upper().endswith(".001") else start_dir

            candidates = [start_dir / "WSIE.DBF", root / "WSIE.DBF"]
            try:
                candidates.extend(root.rglob("WSIE.DBF"))
                candidates.extend(root.rglob("wsie.dbf"))
            except Exception:
                pass

            seen = set()
            for cand in candidates:
                key = str(cand).upper()
                if key in seen or not cand.exists():
                    continue
                seen.add(key)
                try:
                    fields, records = self.read_dbf(str(cand))
                    if records:
                        nazwa = str(records[0].get('NAZWA', '')).strip()
                        if nazwa:
                            return nazwa
                except Exception:
                    continue
        except Exception:
            pass
        return None

    def run_mietek_rozbieznosci_thread(self, mietki_dir_str, excel_dir_str, out_dir_str, bez_nazwisk=False):
        try:
            self.update_status("Porównywanie DBF z Ewidencją i generowanie...", "#0078D7")
            mietki_dir = Path(mietki_dir_str)
            excel_dir = Path(excel_dir_str)
            out_dir = Path(out_dir_str)
            out_dir.mkdir(parents=True, exist_ok=True)

            xls_files = sorted([
                f for f in excel_dir.iterdir()
                if f.is_file() and f.suffix.lower() in {".xls", ".xlsx"} and not f.name.startswith("~$")
            ])
            if not xls_files:
                raise Exception("Brak plików Excel we wskazanym folderze Ewidencji.")

            total = len(xls_files)
            self.start_progress_tracking(total, "Generowanie raportów rozbieżności")
            stat_sukces = 0
            stat_brak_dbf = 0

            def norm_jrej(val):
                try:
                    return str(int(float(str(val).strip())))
                except Exception:
                    return str(val).strip()

            def safe_float(val):
                if not val:
                    return 0.0
                clean_str = str(val).replace('\x00', '').replace(',', '.').strip()
                match = re.search(r'-?\d+(?:\.\d+)?', clean_str)
                if match:
                    try:
                        return float(match.group(0))
                    except Exception:
                        return 0.0
                return 0.0

            def safe_int_str(val):
                if not val:
                    return "0"
                clean_str = str(val).replace('\x00', '').strip()
                match = re.search(r'\d+', clean_str)
                if match:
                    try:
                        return str(int(match.group(0)))
                    except Exception:
                        return "0"
                return "0"

            for idx, path_xls in enumerate(xls_files, start=1):
                self.check_stop()

                nazwa_wsi_fallback = re.sub(r'(?i)_?rozliczone.*$', '', path_xls.stem).strip()
                if not nazwa_wsi_fallback:
                    nazwa_wsi_fallback = path_xls.stem

                self.progress_current_file = nazwa_wsi_fallback
                v_norm = re.sub(r'[\s_\-]', '', nazwa_wsi_fallback.lower())

                # 1. SZUKANIE DBF
                target_dbf = None
                all_dbfs = list(mietki_dir.rglob("D*.DBF")) + list(mietki_dir.rglob("D*.dbf")) + list(
                    mietki_dir.rglob("d*.DBF")) + list(mietki_dir.rglob("d*.dbf"))
                for dbf in all_dbfs:
                    path_norm = re.sub(r'[\s_\-]', '', str(dbf).lower())
                    if v_norm in path_norm:
                        target_dbf = dbf
                        break
                # FALLBACK: jeśli pełna nazwa nie pasuje, spróbuj poszczególnych słów
                # (np. "REJESTR WSZYSCY LESZKOWICE" → spróbuj "leszkowice")
                if not target_dbf:
                    # Bierzemy słowa dłuższe niż 3 znaki, sortujemy od najdłuższego
                    words = sorted(
                        [w for w in nazwa_wsi_fallback.lower().split() if len(w) > 3],
                        key=len, reverse=True
                    )
                    for word in words:
                        word_norm = re.sub(r'[\s_\-]', '', word)
                        for dbf in all_dbfs:
                            path_norm = re.sub(r'[\s_\-]', '', str(dbf).lower())
                            if word_norm in path_norm:
                                target_dbf = dbf
                                break
                        if target_dbf:
                            break
                if not target_dbf and len(all_dbfs) == 1:
                    target_dbf = all_dbfs[0]

                if not target_dbf:
                    self.log(f"  ⚠️ {nazwa_wsi_fallback}: brak dopasowanego pliku D*.DBF (zignorowano).")
                    stat_brak_dbf += 1
                    self.set_progress(idx / total, current_file=nazwa_wsi_fallback, current=idx)
                    continue

                # 1b. NAZWA WSI Z WSIE.DBF (pole NAZWA = 'komórka A2'), fallback: nazwa pliku XLS
                nazwa_wsi = self.read_wsie_nazwa(target_dbf) or nazwa_wsi_fallback
                self.progress_current_file = nazwa_wsi

                try:
                    # 2. ODCZYT DBF (aktualna powierzchnia lasu)
                    fields, records = self.read_dbf(str(target_dbf))
                    dbf_data = []
                    for rec in records:
                        j_rej = safe_int_str(rec.get('NRREJ', ''))
                        nr_dz = str(rec.get('NR_DZIAL', '')).replace('\x00', '').strip()
                        pow_aktualna = safe_float(rec.get('POW', ''))
                        if pow_aktualna > 0:
                            dbf_data.append({'J. rej. norm': j_rej, 'nr_dz': nr_dz, 'pow_aktualna': pow_aktualna})

                    df_dbf = pd.DataFrame(dbf_data)
                    if not df_dbf.empty:
                        df_dbf = df_dbf.groupby(['nr_dz', 'J. rej. norm'], as_index=False)['pow_aktualna'].sum()
                    else:
                        df_dbf = pd.DataFrame(columns=['nr_dz', 'J. rej. norm', 'pow_aktualna'])

                    # 3. ODCZYT EXCELA (Ewidencja)
                    try:
                        tabela_xls, df_full = wczytaj_i_przetworz_wlascicieli(str(path_xls))
                    except Exception as ex:
                        self.log(f"  ❌ Błąd odczytu pliku Excel {path_xls.name}: {ex}")
                        self.set_progress(idx / total, current_file=nazwa_wsi, current=idx)
                        continue

                    tabela_xls['J. rej. norm'] = tabela_xls['J. rej.'].apply(norm_jrej)
                    if not df_full.empty and 'J. rej.' in df_full.columns:
                        df_full['J. rej. norm'] = df_full['J. rej.'].apply(norm_jrej)

                    # 3b. PRZELICZENIE JEDNOSTEK (Excel zawsze w m² -> ha)
                    tabela_xls['pow ls'] = tabela_xls['pow ls'] / 10000.0
                    tabela_xls['pow dz'] = tabela_xls['pow dz'] / 10000.0
                    if not df_full.empty and 'pow dz' in df_full.columns:
                        df_full['pow dz'] = df_full['pow dz'] / 10000.0

                    # 4. ŁĄCZENIE (FULL OUTER JOIN - ubytki i przyrosty)
                    df_merged = pd.merge(df_dbf, tabela_xls, on=['nr_dz', 'J. rej. norm'], how='outer')

                    pelny_wykaz_filtrowany = []
                    for _, row in df_merged.iterrows():
                        nr_dz = str(row.get('nr_dz', '')).strip()
                        j_rej_norm = str(row.get('J. rej. norm', '')).strip()

                        pow_akt = row.get('pow_aktualna', 0.0)
                        if pd.isna(pow_akt):
                            pow_akt = 0.0
                        pow_ewid = row.get('pow ls', 0.0)
                        if pd.isna(pow_ewid):
                            pow_ewid = 0.0

                        diff = round(pow_akt - pow_ewid, 4)
                        if abs(diff) <= 0.0010:
                            continue

                        j_rej = row.get('J. rej.')
                        wlasc = row.get('Właściciel')

                        if (pd.isna(j_rej) or pd.isna(wlasc)) and not df_full.empty:
                            match_full = df_full[
                                (df_full['nr_dz'] == nr_dz) & (df_full['J. rej. norm'] == j_rej_norm)
                                ]
                            if not match_full.empty:
                                if pd.isna(j_rej):
                                    j_rej = match_full.iloc[0]['J. rej.']
                                if pd.isna(wlasc) and 'Właściciel' in match_full.columns:
                                    wlasc = match_full.iloc[0]['Właściciel']

                        if pd.isna(j_rej):
                            j_rej = j_rej_norm
                        if pd.isna(wlasc):
                            wlasc = "Brak danych"

                        # 5. CZYSZCZENIE NAZWISKA DO WYKAZU
                        surowy_wlasciciel = str(wlasc).replace('nan', 'Brak danych')
                        clean_names_list = []
                        blocks = re.split(r'(?m)^(\d+/\d+)\s+\[.*?\]\s*', surowy_wlasciciel)
                        if len(blocks) == 1:
                            blocks = re.split(r'(?m)^(\d+/\d+)\s+\[.*?\]\s*', "1/1 [własność] " + surowy_wlasciciel)

                        for b_idx in range(1, len(blocks), 2):
                            if b_idx + 1 >= len(blocks):
                                break
                            rest = blocks[b_idx + 1]
                            lines = [line.strip() for line in rest.split('\n') if line.strip()]
                            parsing_names = True
                            for line in lines:
                                if line == 'Podmiot grupowy':
                                    continue
                                if parsing_names:
                                    clean_name = re.sub(r'\s*\[(OF|OP|PG)\]', '', line).strip()
                                    if clean_name:
                                        clean_names_list.append(clean_name)
                                    if re.search(r'\[(OF|OP|PG)\]', line):
                                        parsing_names = False

                        if clean_names_list:
                            clean_names = ", ".join(clean_names_list)
                        else:
                            clean_names = surowy_wlasciciel.replace('\n', ' ')

                        pelny_wykaz_filtrowany.append({
                            'J. rej.': j_rej,
                            'właściciel': clean_names,
                            'nr działki': nr_dz,
                            'ls ewidenca': pow_ewid,
                            'przybyło': diff if diff > 0 else None,
                            'ubyło': abs(diff) if diff < 0 else None,
                            'po zmianie': pow_akt
                        })

                    # 5b. SORTOWANIE PO J. REJ. ROSNĄCO (numerycznie)
                    def _sort_key(rec):
                        try:
                            return (0, int(float(str(rec.get('J. rej.', '0')).strip())))
                        except Exception:
                            return (1, str(rec.get('J. rej.', '')))

                    pelny_wykaz_filtrowany.sort(key=_sort_key)

                    # 5c. TRYB BEZ NAZWISK - pusta kolumna właściciela
                    if bez_nazwisk:
                        for rec in pelny_wykaz_filtrowany:
                            rec['właściciel'] = ""

                    # 6. ZAPIS DO EXCELA NA BAZIE SZABLONU
                    if pelny_wykaz_filtrowany:
                        import openpyxl
                        template_path = get_resource_path("BIAŁYNIN KRASÓWKA.xlsx")
                        if Path(template_path).exists():
                            wb_template = openpyxl.load_workbook(template_path)
                            if 'Wykaz rozbieżnosci' in wb_template.sheetnames:
                                ws_rozbieznosci = wb_template['Wykaz rozbieżnosci']
                                ws_rozbieznosci['A3'] = nazwa_wsi.upper()

                                max_r = ws_rozbieznosci.max_row
                                if max_r >= 8:
                                    ws_rozbieznosci.delete_rows(8, max_r - 7)

                                start_row = 8
                                thin_border = Border(
                                    left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'))

                                for i, record in enumerate(pelny_wykaz_filtrowany):
                                    row_idx = start_row + i
                                    ws_rozbieznosci.cell(row=row_idx, column=1, value=record.get('J. rej.', ''))
                                    ws_rozbieznosci.cell(row=row_idx, column=2, value=record.get('właściciel', ''))
                                    ws_rozbieznosci.cell(row=row_idx, column=3, value=record.get('nr działki', ''))
                                    ws_rozbieznosci.cell(row=row_idx, column=4, value=record.get('ls ewidenca', ''))

                                    przybylo_val = record.get('przybyło')
                                    if przybylo_val is not None:
                                        ws_rozbieznosci.cell(row=row_idx, column=5, value=przybylo_val)

                                    ubylo_val = record.get('ubyło')
                                    if ubylo_val is not None:
                                        ws_rozbieznosci.cell(row=row_idx, column=6, value=ubylo_val)

                                    ws_rozbieznosci.cell(row=row_idx, column=7, value=record.get('po zmianie', ''))

                                    for col in range(1, 9):
                                        cell = ws_rozbieznosci.cell(row=row_idx, column=col)
                                        cell.border = thin_border
                                        cell.font = Font(name="Arial", size=10)
                                        if col in [4, 5, 6, 7]:
                                            cell.number_format = '0.0000'

                                suffix = "_WYKAZ ROZBIEZNOSCI_BEZ NAZWISK.xlsx" if bez_nazwisk else "_WYKAZ ROZBIEZNOSCI.xlsx"
                                rozbieznosci_output = out_dir / f"{nazwa_wsi}{suffix}"
                                wb_template.save(str(rozbieznosci_output))
                                self.log(f"  ✅ Utworzono Wykaz Rozbieżności: {rozbieznosci_output.name}")
                                stat_sukces += 1
                            else:
                                self.log(
                                    "  ❌ Szablon 'BIAŁYNIN KRASÓWKA.xlsx' nie zawiera arkusza 'Wykaz rozbieżnosci'!")
                        else:
                            self.log(
                                "  ⚠️ Nie znaleziono pliku wzorcowego 'BIAŁYNIN KRASÓWKA.xlsx'. Upewnij się, że jest w folderze.")
                    else:
                        self.log(f"  ℹ️ Obręb {nazwa_wsi} nie posiada zmian (ubyło/przybyło). Pomijam.")
                        stat_sukces += 1

                except Exception as e:
                    self.log(f"  ❌ Błąd przetwarzania dla '{nazwa_wsi}': {e}")

                self.set_progress(idx / total, current_file=nazwa_wsi, current=idx)

            self.update_status("Zakończono pomyślnie.", "#27ae60", animate=False)
            self.log(f"\n✅ ZAKOŃCZONO. Utworzono lub zweryfikowano: {stat_sukces}/{total}")
            if stat_brak_dbf:
                self.log(f"⚠️ Pomięto {stat_brak_dbf} obrębów z braku bazy D*.DBF.")

            self.after(0, lambda: messagebox.showinfo(
                "Sukces", f"Operacja zakończona (przetworzono {stat_sukces} obrębów)."))

        except InterruptedError:
            self.update_status("Przerwano", "#D83B01", animate=False)
            self.log("\nZADANIE PRZERWANE PRZEZ UŻYTKOWNIKA.")
        except Exception as e:
            self.log(traceback.format_exc())
            self.update_status("Błąd", "#D83B01", animate=False)
        finally:
            self.running = False
            self.after(0, self.restore_all_buttons)

# ==========================================
# GŁÓWNY PUNKT WEJŚCIA PROGRAMU (START)
# ==========================================

    # ZMODYFIKOWANE METODY ZADANIOWE ZWRACAJĄCE LICZNIK DLA DASHBOARDU
